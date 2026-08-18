import os, sys, re, openpyxl, threading, pandas as pd, extract_msg, gc, time, logging, json
from pathlib import Path
from bs4 import BeautifulSoup
import openai
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

logging.basicConfig(
    filename='pipeline_debug.log', 
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='w'
)

STATE_FILE = "pipeline_state.json"

def save_pipeline_state(next_start_idx, folder_path="", status="in_progress"):
    """Saves the next index to process and folder path on disk in real-time."""
    try:
        # Load existing data first so we preserve existing values if not passed
        current = load_pipeline_state()
        saved_folder = folder_path if folder_path else current.get("folder_path", "")
        
        with open(STATE_FILE, "w") as f:
            json.dump({
                "next_start_idx": next_start_idx, 
                "folder_path": saved_folder, 
                "status": status
            }, f)
    except Exception as e:
        logging.error(f"Failed to save state file: {str(e)}")

def load_pipeline_state():
    """Reads saved state on app startup."""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {"next_start_idx": 1, "folder_path": "", "status": "completed"}

# TODO: token must be changed or updated
GITHUB_TOKEN = "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" 

client = openai.OpenAI(
    base_url="https://models.github.ai/inference",  
    api_key=GITHUB_TOKEN
)

def clean_text(val, rm_sig=False):
    if not isinstance(val, str) or not val.strip() or val in ["None", "N/A"]: return "" if rm_sig else val
    val = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F\'"]', '', val).strip()
    val = re.sub(r'\b([A-Z][a-z]+)\s+([A-Z]\.?\s+)?([A-Z][a-z]+)\b', r'\1 \3', val)
    if rm_sig:
        lines = []
        for l in val.splitlines():
            if any(re.search(p, l.lower()) for p in [r'^\s*thank', r'^\s*regard', r'^\s*best', r'^\s*sincer', r'^---']): break
            lines.append(l)
        return "\n".join(lines).strip()
    return val

def format_contact(raw_str, multi=False):
    if multi:
        return "\n".join([format_contact(t.strip()) for t in re.split(r'[;,\n]\s*', raw_str or "") if t.strip()])
    
    contact = clean_text(raw_str)
    if not contact or contact.lower() in ["email", "unknown", "unknown sender"]: return "Unknown Contact"
    
    em_match = re.search(r'<([^>]+)>', contact)
    email = em_match.group(1).strip().lower() if em_match else ("" if " " in contact or "@" not in contact else contact)
    
    name = re.sub(r'<[^>]+>', '', contact).strip() if email else contact.split("@")[0]
    
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        name = f"{parts[1]} {parts[0]}" if len(parts) == 2 else name
        
    name = " ".join([w.capitalize() for w in clean_text(name).split()])
    
    if not email:
        return name
            
    return f"{name} <{email}>"

def get_ai_summary(body, subject="", force_skip_api=False):
    cleaned = clean_text(body, rm_sig=True)
    subj_lower = subject.lower()
    
    if not cleaned: 
        return "Empty Body", 0
        
    # skip system notifications, calendar invites, and auto-replies (0 tokens)
    if any(k in subj_lower for k in ["automatic reply:", "out of office:", "read:", "accepted:", "declined:"]):
        return (cleaned if len(cleaned) < 150 else cleaned[:150] + "..."), 0
        
    if "when:" in cleaned.lower() and "where:" in cleaned.lower():
        return "Meeting / Calendar Invitation details omitted.", 0

    # skip short emails by character length or word count (0 tokens)
    # adjust 300 or 40 to be higher or lower
    words = cleaned.split()
    if len(cleaned) < 300 or len(words) < 40:
        return cleaned, 0

    # expanded short-phrase or quick reply detection (0 tokens)
    short_phrases = [
        r'^\s*thank', r'^\s*got it', r'^\s*see attached', r'^\s*will do', 
        r'^\s*ok', r'^\s*okay', r'^\s*approved', r'^\s*please see', 
        r'^\s*sounds good', r'^\s*fyi', r'^\s*per our conversation'
    ]
    if any(re.search(p, cleaned.lower()) for p in short_phrases):
        return cleaned, 0

    if force_skip_api:
        return "AI Summary Skipped (API Limit Hit)", 0

    est_input_tokens = len(cleaned[:3500]) // 4

    try:
        prompt = (
            f"Summarize the email body in 1-4 SHORT sentences. Never have redundant information in your response. "
            f"High-speed facts only. No intros like 'This email discusses'. Never mention greetings, never include middle and last names unless absolutely necessary, and never include bullet points. "
            f"Never put your response into markdown format. Your responses should only be plain text. "
            f"Never include closings or any sentence related to 'contact me with any questions'. "
            f"Just output the plain text summary straight.\n\n{cleaned[:3500]}"
        )
        
        response = client.chat.completions.create(
            model="openai/gpt-4o-mini",
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0.0,
            max_tokens=150
        )
        
        out = response.choices[0].message.content.strip()
        est_output_tokens = len(out) // 4
        return clean_text(out), (est_input_tokens + est_output_tokens)
        
    except openai.RateLimitError:
        return "API_LIMIT_TRIGGERED", est_input_tokens
    except Exception as ai_err: 
        logging.warning(f"GitHub Models API generation dropped: {str(ai_err)}")
        err_msg = str(ai_err).lower()
        status_code = getattr(ai_err, 'status_code', None)
        if status_code == 429 or any(k in err_msg for k in ["429", "rate limit", "quota", "resource_exhausted", "too many requests"]):
            return "API_LIMIT_TRIGGERED", est_input_tokens
        return "AI Summary Skipped (API Error)", est_input_tokens

extracted_data_buffer = []  
full_body_cache = {} 
processed_filenames = set()
api_limit_hit = False
reset_time_str = ""
export_menu_enabled = False 
pipeline_running = False
tree_item_ids = []

MAX_TOKENS_PER_MINUTE = 150000
tokens_used_this_minute = 0
minute_window_start = time.time()

def run_pipeline(folder_path, limit_str, start_pos_str, ui_widgets):
    global extracted_data_buffer, api_limit_hit, reset_time_str, export_menu_enabled, tree_item_ids, pipeline_running
    global tokens_used_this_minute, minute_window_start
    
    lbl_file, lbl_status, lbl_time, lbl_tokens, pbar, btn_start, btn_export, tree = ui_widgets
    
    pipeline_running = True
    extracted_data_buffer.clear()
    tree_item_ids.clear()
    full_body_cache.clear()
    processed_filenames.clear()
    
    tokens_used_this_minute = 0
    minute_window_start = time.time()
    
    api_limit_hit = False
    first_failed_index = None
    reset_time_str = ""
    export_menu_enabled = False
    update_dropdown_menu_options(has_failed_items=False)
    
    for item in tree.get_children():
        tree.delete(item)
        
    if not GITHUB_TOKEN or "your_github_personal" in GITHUB_TOKEN:
        messagebox.showerror("Configuration Error", "Please paste your GitHub Token into the code.")
        set_button_state(btn_start, "normal", "#1F4E78")
        pipeline_running = False
        return
    
    if not folder_path:
        messagebox.showerror("Configuration Error", "Please select a source folder first.")
        set_button_state(btn_start, "normal", "#1F4E78")
        pipeline_running = False
        return

    start_idx = 1
    if start_pos_str:
        try:
            start_idx = int(start_pos_str)
            if start_idx <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Configuration Error", "Start index must be a number greater than 0.")
            set_button_state(btn_start, "normal", "#1F4E78")
            pipeline_running = False
            return

    max_emails = None
    if limit_str:
        try:
            max_emails = int(limit_str)
            if max_emails <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Configuration Error", "Limit must be a number greater than 0.")
            set_button_state(btn_start, "normal", "#1F4E78")
            pipeline_running = False
            return

    if not os.path.exists(folder_path):
        messagebox.showerror("Directory Error", "The selected folder path does not exist.")
        set_button_state(btn_start, "normal", "#1F4E78")
        pipeline_running = False
        return

    lbl_status.config(text="Status: Searching folder for files...", fg="#1565C0")
    root.update_idletasks()

    base_dir = Path(folder_path)
    msg_files = [str(p) for p in base_dir.rglob('*.msg')]
    total_files = len(msg_files)
    active_files = msg_files[start_idx - 1:]
    
    run_total = max_emails if (max_emails is not None and max_emails < len(active_files)) else len(active_files)

    if total_files == 0 or len(active_files) == 0:
        pbar['value'] = 100
        lbl_status.config(text="Status: Complete. No email files found.", fg="#1565C0")
        set_button_state(btn_start, "normal", "#1F4E78")
        pipeline_running = False
        return

    pbar['value'] = 0
    pbar['maximum'] = run_total
    start_time = time.time()
    has_any_skips = False

    for idx, file_path in enumerate(active_files, 1):
        if idx > run_total: break

        fname = os.path.basename(file_path)
        current_overall_idx = (start_idx - 1) + idx

        if fname in processed_filenames:
            logging.info(f"Skipping duplicate filename: {fname}")
            continue
        processed_filenames.add(fname)

        lbl_file.config(text=f"Current File: {fname[:75]}..." if len(fname) > 78 else f"Current File: {fname}")
        
        now = time.time()
        if now - minute_window_start >= 60:
            tokens_used_this_minute = 0
            minute_window_start = now
            
        elapsed_sec = time.time() - start_time
        avg_time = elapsed_sec / idx if idx > 0 else 1
        rem_sec = max(0.0, (run_total - idx) * avg_time)
        
        try:
            lbl_status.config(text=f"Status: Processing email {idx} of {run_total}", fg="#333333")
            root.update_idletasks()
            
            with extract_msg.Message(file_path) as msg:
                raw_subj = msg.subject if msg.subject else "No Subject"
                raw_sender = format_contact(msg.sender) if msg.sender else "Unknown Sender"
                raw_date = msg.date.strftime("%A, %B %d, %Y") if msg.date else "Missing Date"
                raw_time = msg.date.strftime("%I:%M:%S %p") if msg.date else "Missing Time"
                raw_to = format_contact(msg.to, multi=True) if msg.to else "N/A"
                raw_cc = format_contact(msg.cc, multi=True) if msg.cc else ""
                raw_attachments = "\n".join([a.longFilename for a in msg.attachments if a.longFilename]) or "None"
                
                body = msg.body if hasattr(msg, 'body') and msg.body else ""
                if not body and hasattr(msg, 'htmlBody') and msg.htmlBody:
                    body = BeautifulSoup(msg.htmlBody, 'html.parser').get_text(' ')
                
                ai_sum, tokens_used = get_ai_summary(body, subject=raw_subj, force_skip_api=api_limit_hit)
                
                if ai_sum == "API_LIMIT_TRIGGERED":
                    api_limit_hit = True
                    first_failed_index = current_overall_idx
                    reset_dt = datetime.now() + timedelta(days=1)
                    reset_time_str = reset_dt.strftime("%m/%d at %I:%M %p")
                    
                    # before API call in loop:
                    save_pipeline_state(first_failed_index, folder_path=folder_path, status="api_limit_hit")

                    # at completion:
                    save_pipeline_state(1, folder_path=folder_path, status="completed")
                    
                    # halt further API requests immediately
                    break

                tokens_used_this_minute += tokens_used
                lbl_time.config(text=f"Time Elapsed: {int(elapsed_sec)}s  |  Remaining: {int(rem_sec)}s")
                available_tokens = max(0, MAX_TOKENS_PER_MINUTE - tokens_used_this_minute)
                lbl_tokens.config(text=f"Available Tokens (This Minute): {available_tokens:,} / 150,000")

                if "skipped" in ai_sum.lower() or "error" in ai_sum.lower():
                    has_any_skips = True

                row_vals = [fname, raw_subj, raw_sender, raw_to, raw_cc, raw_attachments, raw_date, raw_time, ai_sum]
                extracted_data_buffer.append(row_vals)
                full_body_cache[fname] = body if body.strip() else "No body text content available."
                
                ui_row = ["☐"] + [clean_text(str(v)).replace('\n', ' ') for v in row_vals]
                item_id = tree.insert("", "end", values=ui_row)
                tree_item_ids.append(item_id)
                tree.yview_moveto(1.0) 
                
            gc.collect()
        except Exception as loop_err:
            logging.error(f"Error processing file {fname}: {str(loop_err)}")
            has_any_skips = True
            row_vals = [fname, "Error", "Error", "N/A", "", "None", "N/A", "N/A", f"Error: {str(loop_err)}"]
            extracted_data_buffer.append(row_vals)
            full_body_cache[fname] = f"Error reading source body: {str(loop_err)}"
            
            ui_row = ["☐"] + [v.replace('\n', ' ') for v in row_vals]
            item_id = tree.insert("", "end", values=ui_row)
            tree_item_ids.append(item_id)

        time.sleep(4.1)
        pbar['value'] = idx
        root.update_idletasks()

    actual_processed_count = len(extracted_data_buffer)

    if api_limit_hit and first_failed_index is not None:
        pbar['value'] = actual_processed_count
        pbar['maximum'] = actual_processed_count
        
        lbl_status.config(
            text=f"Status: Stopped at item {actual_processed_count} (API Limit Hit. Resets: {reset_time_str})", 
            fg="#E65100"
        )
        
        # update Start Index to exact resume point
        ent_start_idx.delete(0, tk.END)
        ent_start_idx.insert(0, str(first_failed_index))
        
        # subtract processed items from total limit input field
        if max_emails is not None:
            remaining_limit = max(0, max_emails - actual_processed_count)
            ent_limit.delete(0, tk.END)
            ent_limit.insert(0, str(remaining_limit))
        
        messagebox.showwarning(
            "API Limit Reached", 
            f"Process stopped early after processing {actual_processed_count} email(s).\n\n"
            f"• Next Start Index updated to: {first_failed_index}\n"
            f"• Remaining Limit updated to: {ent_limit.get()}\n"
            f"• Token quota resets around: {reset_time_str}."
        )
    else:
        pbar['value'] = run_total
        save_pipeline_state(1, status="completed")
        lbl_status.config(text=f"Status: Process Complete. Loaded {actual_processed_count} total records.", fg="#2E7D32")
        messagebox.showinfo("Success", f"All {actual_processed_count} files loaded successfully.")
        
    set_button_state(btn_start, "normal", "#1F4E78")
    set_button_state(btn_export, "normal", "#2E7D32")
    export_menu_enabled = True
    update_dropdown_menu_options(has_failed_items=has_any_skips)
    pipeline_running = False

def handle_row_click(event):
    item = tree.identify_row(event.y)
    if not item: return
    
    col = tree.identify_column(event.x)
    if col != "#1": return 
    
    current_values = list(tree.item(item, "values"))
    if not current_values: return
    
    if current_values[0] == "☐":
        current_values[0] = "☑"
    else:
        current_values[0] = "☐"
        
    tree.item(item, values=current_values)

def toggle_all_rows(select_all=True):
    symbol = "☑" if select_all else "☐"
    for item_id in tree_item_ids:
        vals = list(tree.item(item_id, "values"))
        if vals:
            vals[0] = symbol
            tree.item(item_id, values=vals)

def invert_row_selection():
    """Flips every row's check state (checked -> unchecked, unchecked -> checked)."""
    for item_id in tree_item_ids:
        vals = list(tree.item(item_id, "values"))
        if vals:
            vals[0] = "☐" if vals[0] == "☑" else "☑"
            tree.item(item_id, values=vals)

def open_full_body_window(event):
    item = tree.identify_row(event.y)
    if not item: return
    
    vals = tree.item(item, "values")
    if not vals or len(vals) < 2: return
    
    filename = vals[1] 
    subject = vals[2]
    
    body_text = full_body_cache.get(filename, "No source body matching cache index available.")
    
    pop = tk.Toplevel(root)
    pop.title(f"Full Message Viewer - {filename}")
    pop.geometry("680x520")
    pop.transient(root)
    
    tk.Label(pop, text=f"Subject: {subject}", font=("Calibri", 11, "bold"), anchor="w", justify="left", padx=10, pady=6).pack(fill="x")
    
    txt_area = scrolledtext.ScrolledText(pop, wrap="word", font=("Calibri", 10), padx=8, pady=8)
    txt_area.insert(tk.END, body_text)
    txt_area.config(state="disabled")
    txt_area.pack(fill="both", expand=True, padx=10, pady=(0, 10))

def save_to_excel_file():
    if not extracted_data_buffer:
        messagebox.showwarning("Empty Collection", "There is no extracted data to write yet.")
        return
        
    output_excel = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if not output_excel: return 

    cols = ["FileName", "Subject", "Sender", "Recipient", "CC", "Attachments", "Date", "Time", "Summary"]
    
    try:
        wb = openpyxl.load_workbook(output_excel) if os.path.exists(output_excel) else openpyxl.Workbook()
        ws = wb.active
        
        existing_records = {}
        if os.path.exists(output_excel):
            for row_num in range(2, ws.max_row + 1):
                f_cell = str(ws.cell(row=row_num, column=1).value or "").strip()
                s_cell = str(ws.cell(row=row_num, column=9).value or "").strip()
                if f_cell: existing_records[f_cell] = {"row": row_num, "summary": s_cell}
        else:
            for c_idx, name in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c_idx, value=name)
                cell.fill, cell.font, cell.alignment = PatternFill("solid", "1F4E78"), Font("Calibri", 11, bold=True, color="FFFFFF"), Alignment("left", "center")
            ws.row_dimensions[1].height = 26

        styles = {"font": Font("Calibri", 11), "border": Border(*[Side(style="thin", color="D9D9D9")]*4), "align": Alignment("left", "top", wrap_text=True)}

        for row_vals in extracted_data_buffer:
            fname = row_vals[0]
            
            if fname in existing_records:
                curr_sum = existing_records[fname]["summary"]
                if curr_sum and "skipped" not in curr_sum.lower() and "failed" not in curr_sum.lower():
                    continue 
                target_row = existing_records[fname]["row"]
            else:
                target_row = ws.max_row + 1

            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=target_row, column=c_idx, value=clean_text(val))
                cell.font, cell.border, cell.alignment = styles["font"], styles["border"], styles["align"]

        for r in range(ws.max_row, max(1, ws.max_row - 5), -1):
            if "NOTICE:" in str(ws.cell(row=r, column=1).value or ""): ws.delete_rows(r)

        if api_limit_hit:
            nr = ws.max_row + 2
            ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=9)
            warning_cell = ws.cell(row=nr, column=1)
            warning_cell.value = f"NOTICE: API Limit reached. Remaining summaries skipped, try again later."
            warning_cell.font = Font("Calibri", 11, bold=True, color="9C0006")
            warning_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            warning_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[nr].height = 24

        widths = {'A': 25, 'B': 30, 'C': 25, 'D': 30, 'E': 30, 'F': 25, 'G': 25, 'H': 15, 'I': 65}
        for col, w in widths.items(): ws.column_dimensions[col].width = w
        
        wb.save(output_excel)
        messagebox.showinfo("Export Success", f"Successfully saved file to:\n{os.path.basename(output_excel)}")
    except PermissionError:
        messagebox.showerror("Write Lock Error", "Failed to save updates. Close the file in Excel and try again.")

def copy_selection_to_clipboard(copy_all=False):
    if not extracted_data_buffer:
        messagebox.showwarning("Empty Collection", "No data available to copy.")
        return

    lines = []
    headers = ["FileName", "Subject", "Sender", "Recipient", "CC", "Attachments", "Date", "Time", "Summary"]
    
    if include_headers_var.get():
        lines.append("\t".join(headers))
    
    has_checked_items = False
    
    for item_id in tree_item_ids:
        vals = list(tree.item(item_id, "values"))
        if not vals: continue
        
        is_checked = (vals[0] == "☑")
        data_vals = vals[1:] 
        
        if copy_all or (is_checked and not copy_all):
            if is_checked: has_checked_items = True
            lines.append("\t".join([str(v) for v in data_vals]))

    if not copy_all and not has_checked_items:
        for item_id in tree_item_ids:
            vals = list(tree.item(item_id, "values"))[1:]
            lines.append("\t".join([str(v) for v in vals]))
        msg_title, msg_desc = "Copied Data", f"No selections checked. Defaulted to copying all processed entries."
    elif copy_all:
        msg_title, msg_desc = "Copied Data", f"Successfully copied all processed entries to clipboard."
    else:
        msg_title, msg_desc = "Copied Data", f"Successfully copied checked row(s) to clipboard."

    clipboard_text = "\n".join(lines)
    root.clipboard_clear()
    root.clipboard_append(clipboard_text)
    root.update()
    
    messagebox.showinfo(msg_title, msg_desc)

def select_completed_summaries_only():
    for item_id in tree_item_ids:
        vals = list(tree.item(item_id, "values"))
        if not vals: continue
        
        summary_text = str(vals[9]).lower() 
        if "skipped" not in summary_text and "error" not in summary_text:
            vals[0] = "☑"
        else:
            vals[0] = "☐"
        tree.item(item_id, values=vals)

def post_export_dropdown():
    if not export_menu_enabled:
        return
    x = btn_export.winfo_rootx()
    y = btn_export.winfo_rooty() + btn_export.winfo_height()
    export_menu.post(x, y)

def update_dropdown_menu_options(has_failed_items=False):
    export_menu.delete(0, tk.END)
    
    excel_menu = tk.Menu(export_menu, tearoff=0, font=("Calibri", 10))
    excel_menu.add_command(label="Save Entire Set to Excel File (.xlsx)", command=save_to_excel_file)
    export_menu.add_cascade(label="Excel Document Exports", menu=excel_menu)
    export_menu.add_separator()
    
    clip_menu = tk.Menu(export_menu, tearoff=0, font=("Calibri", 10))
    clip_menu.add_command(label="Copy Checked Row(s) to Clipboard", command=lambda: copy_selection_to_clipboard(copy_all=False))
    clip_menu.add_command(label="Copy Entire Feed Table to Clipboard", command=lambda: copy_selection_to_clipboard(copy_all=True))
    clip_menu.add_separator()
    clip_menu.add_checkbutton(label="Include Field Headers on Copy", variable=include_headers_var)
    export_menu.add_cascade(label="Clipboard Operations", menu=clip_menu)
    export_menu.add_separator()
    
    filter_menu = tk.Menu(export_menu, tearoff=0, font=("Calibri", 10))
    filter_menu.add_command(label="Select Only Completed Summaries", command=select_completed_summaries_only)
    filter_menu.add_command(label="Select All Rows", command=lambda: toggle_all_rows(select_all=True))
    filter_menu.add_command(label="Clear All Selections", command=lambda: toggle_all_rows(select_all=False))
    filter_menu.add_command(label="Invert Current Selection", command=invert_row_selection)
    export_menu.add_cascade(label="Selection Rules and Filters", menu=filter_menu)

def on_closing_handler():
    if pipeline_running:
        if messagebox.askyesno("Process Running", "An email process is running. Force quitting may cause file data logging loss. Exit anyway?"):
            root.destroy()
    else:
        root.destroy()

def show_user_guide():
    guide_text = (
        "Instructions:\n\n"
        "1. Select Folder\n"
        "   Click 'Browse' and pick the folder where your emails are stored.\n\n"
        "2. Process Emails\n"
        "   Click 'START' to process the folder.\n\n"
        "3. Selecting Specific Rows\n"
        "   Click directly inside the 'Select' column checkbox slot to toggle selections manually. "
        "Or use the 'Select All' / 'Clear All' quick actions directly above the table frame.\n\n"
        "4. View Entire Message Text\n"
        "   Double-click any entry row across the feed list grid to read its full message text layout body.\n\n"
        "5. Save or Copy Data\n"
        "   Click 'EXPORT LOG DATA' to access clean subcategories for Excel Saving and Clipboard management rules."
    )
    messagebox.showinfo("User Guide", guide_text)

def set_button_state(btn, state, active_color):
    if state == "disabled":
        btn.config(state=tk.DISABLED, bg="#D3D3D3", activebackground="#D3D3D3")
    else:
        btn.config(state=tk.NORMAL, bg=active_color, activebackground=active_color)

def start_thread():
    set_button_state(btn_start, "disabled", "#1F4E78")
    set_button_state(btn_export, "disabled", "#2E7D32")
    folder = ent_folder.get().strip()
    limit = ent_limit.get().strip()
    start_pos = ent_start_idx.get().strip()
    
    t = threading.Thread(target=run_pipeline, args=(folder, limit, start_pos, (lbl_curr_file, lbl_status, lbl_time, lbl_tokens, pbar, btn_start, btn_export, tree)))
    t.daemon = True
    t.start()

def browse_folder():
    res = filedialog.askdirectory()
    if res: 
        norm_path = os.path.normpath(res)
        ent_folder.delete(0, tk.END)
        ent_folder.insert(0, norm_path)
        # save path immediately upon selection
        current_state = load_pipeline_state()
        save_pipeline_state(current_state.get("next_start_idx", 1), folder_path=norm_path, status=current_state.get("status", "completed"))

root = tk.Tk()
root.title("Email Summarizer")
root.geometry("1060x710")
root.protocol("WM_DELETE_WINDOW", on_closing_handler)

include_headers_var = tk.BooleanVar(value=False)

frame = tk.LabelFrame(root, text=" Settings ", padx=15, pady=10, font=("Calibri", 11, "bold"))
frame.pack(padx=15, pady=10, fill="x")

# Source Folder
tk.Label(frame, text="Source Folder containing emails (.msg):", font=("Calibri", 10)).grid(row=0, column=0, sticky="w", pady=2)
ent_folder = tk.Entry(frame, width=92, font=("Calibri", 10))
ent_folder.grid(row=1, column=0, padx=(0, 5), pady=2)
tk.Button(frame, text="Browse...", command=browse_folder, width=12).grid(row=1, column=1, pady=2)

canvas_help = tk.Canvas(frame, width=26, height=26, highlightthickness=0)
canvas_help.grid(row=0, column=1, sticky="e", pady=2)
oval_id = canvas_help.create_oval(2, 2, 24, 24, fill="#F0F0F0", outline="#B0B0B0")
text_id = canvas_help.create_text(13, 13, text="?", font=("Calibri", 10, "bold"), fill="#333333")

canvas_help.tag_bind(oval_id, "<Button-1>", lambda e: show_user_guide())
canvas_help.tag_bind(text_id, "<Button-1>", lambda e: show_user_guide())
canvas_help.tag_bind(oval_id, "<Enter>", lambda e: canvas_help.itemconfig(oval_id, fill="#E0E0E0"))
canvas_help.tag_bind(oval_id, "<Leave>", lambda e: canvas_help.itemconfig(oval_id, fill="#F0F0F0"))

# limit & Start Index
limit_frame = tk.Frame(frame)
limit_frame.grid(row=2, column=0, columnspan=2, sticky="w", pady=(8,2))

tk.Label(limit_frame, text="Start Index:", font=("Calibri", 10)).pack(side="left")
ent_start_idx = tk.Entry(limit_frame, width=8, font=("Calibri", 10))
ent_start_idx.pack(side="left", padx=5)

tk.Label(limit_frame, text="Limit total items:", font=("Calibri", 10)).pack(side="left", padx=(15, 0))
ent_limit = tk.Entry(limit_frame, width=8, font=("Calibri", 10))
ent_limit.pack(side="left", padx=5)

saved_state = load_pipeline_state()

# load saved folder path
saved_folder_path = saved_state.get("folder_path", "")
if saved_folder_path:
    ent_folder.insert(0, saved_folder_path)

# load saved start index
resume_idx = saved_state.get("next_start_idx", 1)
ent_start_idx.insert(0, str(resume_idx))

canvas_help = tk.Canvas(frame, width=26, height=26, highlightthickness=0)
canvas_help.grid(row=0, column=1, sticky="e", pady=2)
oval_id = canvas_help.create_oval(2, 2, 24, 24, fill="#F0F0F0", outline="#B0B0B0")
text_id = canvas_help.create_text(13, 13, text="?", font=("Calibri", 10, "bold"), fill="#333333")

canvas_help.tag_bind(oval_id, "<Button-1>", lambda e: show_user_guide())
canvas_help.tag_bind(text_id, "<Button-1>", lambda e: show_user_guide())
canvas_help.tag_bind(oval_id, "<Enter>", lambda e: canvas_help.itemconfig(oval_id, fill="#E0E0E0"))
canvas_help.tag_bind(oval_id, "<Leave>", lambda e: canvas_help.itemconfig(oval_id, fill="#F0F0F0"))

limit_frame = tk.Frame(frame)
limit_frame.grid(row=2, column=0, columnspan=2, sticky="w", pady=(8,2))

tk.Label(limit_frame, text="Start Index:", font=("Calibri", 10)).pack(side="left")
ent_start_idx = tk.Entry(limit_frame, width=8, font=("Calibri", 10))
ent_start_idx.pack(side="left", padx=5)

saved_state = load_pipeline_state()
resume_idx = saved_state.get("next_start_idx", 1)
ent_start_idx.insert(0, str(resume_idx))

tk.Label(limit_frame, text="Limit total items:", font=("Calibri", 10)).pack(side="left", padx=(15, 0))
ent_limit = tk.Entry(limit_frame, width=8, font=("Calibri", 10))
ent_limit.pack(side="left", padx=5)

tree_frame = tk.LabelFrame(root, text=" Output Feed ", padx=10, pady=10, font=("Calibri", 11, "bold"))
tree_frame.pack(padx=15, pady=5, fill="both", expand=True)

action_row = tk.Frame(tree_frame, pady=2)
action_row.pack(fill="x", side="top", anchor="w")
tk.Button(action_row, text="Select All Rows", font=("Calibri", 9), bg="#F5F5F5", bd=1, relief="groove", command=lambda: toggle_all_rows(select_all=True), padx=6).pack(side="left", padx=(0, 6))
tk.Button(action_row, text="Clear All Selections", font=("Calibri", 9), bg="#F5F5F5", bd=1, relief="groove", command=lambda: toggle_all_rows(select_all=False), padx=6).pack(side="left")

tk.Button(action_row, text="Invert Selection", font=("Calibri", 9), bg="#F5F5F5", bd=1, relief="groove", command=invert_row_selection, padx=6).pack(side="left", padx=(6, 0))

tk.Label(action_row, text="💡 Double-click any row item to open its full email text window view.", font=("Calibri", 9), fg="#666666").pack(side="right", padx=5)

cols_headers = ["Select", "FileName", "Subject", "Sender", "Recipient", "CC", "Attachments", "Date", "Time", "Summary"]
tree = ttk.Treeview(tree_frame, columns=cols_headers, show="headings", selectmode="browse")

style = ttk.Style()
style.configure("Treeview", rowheight=20)

vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

widths = [50, 130, 130, 100, 100, 60, 80, 80, 60, 240]
for col, width in zip(cols_headers, widths):
    tree.heading(col, text=col, anchor="w")
    tree.column(col, width=width, anchor="w", stretch=True if col != "Select" else False)

tree.pack(fill="both", expand=True, side="left")
vsb.pack(fill="y", side="right")
hsb.pack(fill="x", side="bottom", before=tree)

tree.bind("<ButtonRelease-1>", handle_row_click)
tree.bind("<Double-1>", open_full_body_window)

bottom_frame = tk.Frame(root, padx=15, pady=5)
bottom_frame.pack(fill="x", side="bottom")

lbl_curr_file = tk.Label(bottom_frame, text="Current File: Ready", font=("Calibri", 10, "italic"), anchor="w")
lbl_curr_file.pack(fill="x", pady=1)

lbl_status = tk.Label(bottom_frame, text="Status: Awaiting configuration.", font=("Calibri", 10), anchor="w")
lbl_status.pack(fill="x", pady=1)

lbl_time = tk.Label(bottom_frame, text="Time Elapsed: 0s  |  Remaining: 0s", font=("Calibri", 10), anchor="w")
lbl_time.pack(fill="x", pady=1)

lbl_tokens = tk.Label(bottom_frame, text="Available Tokens (This Minute): 150,000 / 150,000", font=("Calibri", 10), anchor="w")
lbl_tokens.pack(fill="x", pady=1)

pbar = ttk.Progressbar(bottom_frame, orient="horizontal", mode="determinate")
pbar.pack(fill="x", pady=(5, 10))

btn_frame = tk.Frame(bottom_frame)
btn_frame.pack(fill="x", pady=5)

btn_start = tk.Button(btn_frame, text="START", bg="#1F4E78", fg="white", font=("Calibri", 10, "bold"), pady=6, command=start_thread, width=32, bd=0, relief="flat")
btn_start.pack(side="left", padx=(0, 10))

btn_export = tk.Button(btn_frame, text="EXPORT LOG DATA  ▾", bg="#D3D3D3", fg="white", font=("Calibri", 10, "bold"), pady=6, command=post_export_dropdown, width=32, bd=0, relief="flat", state=tk.DISABLED)
btn_export.pack(side="left")

export_menu = tk.Menu(root, tearoff=0, font=("Calibri", 10))
update_dropdown_menu_options(has_failed_items=False)

root.mainloop()